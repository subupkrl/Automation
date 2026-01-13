import openpyxl

def dataGenerator():
    workbook=openpyxl.load_workbook("../DataGenerator/data/registerdata.xlsx")
    sh=workbook['Sheet1']
    row=sh.max_row
    li=[]
    li1=[]
    for i in range(1,row+1):
        li1=[]
        fname=sh.cell(i,1)
        lname=sh.cell(i,2)
        month=sh.cell(i,3)
        day=sh.cell(i,4)
        email=sh.cell(i,5)
        pwd=sh.cell(i,6)
        gender=sh.cell(i,7)
        year=sh.cell(i,8)
        li1.insert(0,fname.value)
        li1.insert(1,lname.value)
        li1.insert(2,month.value)
        li1.insert(3,day.value)
        li1.insert(4,year.value)
        li1.insert(5,email.value)
        li1.insert(6,pwd.value)
        li1.insert(7,gender.value)
        li.insert(i-1,li1)
    print(li)
    return li
    
